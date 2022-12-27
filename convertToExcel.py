# -*- coding: utf-8 -*-
"""
Created on Tue Dec 27 10:00:19 2022


@author: Azmi Deliaslan
"""
from database import Database
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

def convert():
    # Create workbook
    wb = openpyxl.Workbook()
    # Get workbook active sheet
    sheet = wb.active
    # Creating Headings
    sheet.cell(row=1, column=1).value = "Product ID"
    sheet.cell(row=1, column=2).value = "Product Category"
    sheet.cell(row=1, column=3).value = "Product Brand"
    sheet.cell(row=1, column=4).value = "Product Name"
    sheet.cell(row=1, column=5).value = "Product Stock"
    sheet.cell(row=1, column=6).value = "Cost Price"
    sheet.cell(row=1, column=7).value = "Selling Price"

    # Styling Headings
    for i in range(1, 8):
        sheet.cell(row=1, column=i).font = Font(bold=True, name="arial", size=15)
        sheet.cell(row=1, column=i).fill = PatternFill(patternType="solid", fgColor="00e5ee")
        sheet.cell(row=1, column=i).alignment = Alignment(horizontal="center")

    # Setting column dimensions
    for col in "ABCDEFGH":
        sheet.column_dimensions[col].width = 25

    # Retrieving date from database
    db = Database("products.db")
    rows = db.fetch_all_rows()

    # Filling date from Database into the worksheet
    for sheet_row, row in enumerate(rows):
        for sheet_col, item in enumerate(row):
            sheet.cell(row=sheet_row + 2, column=sheet_col + 1).value = item
            sheet.cell(row=sheet_row + 2, column=sheet_col + 1).alignment = Alignment(horizontal="center")

    wb.save("product_list.xlsx")


def calc_profit():
    # loading excel file to read and write to
    wb = openpyxl.load_workbook(filename="product_list.xlsx")

    sheet = wb.active

    # Creating  and styling Profit Column Heading
    sheet.cell(row=1, column=8).value = "Profit"
    sheet.cell(row=1, column=8).font = Font(bold=True, name="arial", size=15)
    sheet.cell(row=1, column=8).fill = PatternFill(patternType="solid", fgColor="00e5ee")
    sheet.cell(row=1, column=8).alignment = Alignment(horizontal="center")

    # Calculating Profit
    row_index = 2
    for cell in sheet["H"][1:]:
        cell.value = f"=G{row_index}-F{row_index}"
        row_index = row_index + 1
        cell.alignment = Alignment(horizontal="center")

    # Adding Conditional Formatting on entire row
    yellow_background = PatternFill(bgColor="e9ee9e")
    diff_style = DifferentialStyle(fill=yellow_background)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$H1<0"]
    sheet.conditional_formatting.add(f"A1:H{sheet.max_column}", rule)

    # Same formula as but for a cell in a column
    # sheet.conditional_formatting.add(f'H2:H{sheet.max_column}', CellIsRule(operator='lessThan', formula=['0'], fill=yellow_background))
    wb.save("product_list.xlsx")



